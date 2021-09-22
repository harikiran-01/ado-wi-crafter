from openpyxl import load_workbook


def get_all_rows(excel_file_path):
    workbook = load_workbook(filename=excel_file_path, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    workbook.close()
    return row_iter
